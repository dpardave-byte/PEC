from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_XLSX = Path(
    r"C:/Users/Dpard/OneDrive/Documentos/DGPPCS/BM/Visor/Sistema Seguimiento-PEC - 20.04.26.xlsx"
)
OUTPUT_JSON = Path(r"C:/Users/Dpard/OneDrive/Escritorio/EC/pec_tracking_data.json")
OUTPUT_JS = Path(r"C:/Users/Dpard/OneDrive/Escritorio/EC/pec_tracking_data.js")
SHEET_NAME = "CRONOGRAMA"
HEADER_ROW = 7
TODAY = date(2026, 4, 25)
SELECTED_COLUMNS = [1, 2, 3, 4, 5, 6, 7, 8, 248, 249, 250]
CANONICAL_ROOT_3_TITLE = "Plazo para cumplimiento de condiciones de efectividad"
CANONICAL_ROOT_4_TITLE = "Inicio de Efectividad del Préstamo"
CANONICAL_ETGP_TITLE = "Conformación Equipo ETGP - (DGPPCS)"
CANONICAL_EDT_BY_ID = {
    "case-045": "4.1",
    "case-046": "4.1.1",
    "case-047": "4.1.2",
    "case-048": "4.1.3",
    "case-078": "",
    "case-079": "4",
}


def to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def to_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def to_number(value: Any) -> float | int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text.upper() == "#REF!":
        return None
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return None


def normalize_header(value: Any, col_idx: int) -> str:
    text = to_text(value)
    if not text:
        return f"col_{col_idx}"
    text = re.sub(r"\s+", " ", text)
    return text


def clean_status(value: Any) -> str:
    text = to_text(value)
    if not text:
        return "Sin estado"
    return text


def derive_progress(status: str) -> int:
    status_lower = status.lower()
    if "completado" in status_lower:
        return 100
    if "proceso" in status_lower:
        return 55
    if "pendiente" in status_lower:
        return 0
    return 15


def derive_alert(start_date: str, end_date: str, status: str) -> str:
    if "completado" in status.lower():
        return "ok"
    if not end_date:
        return "sin_fecha"
    end_dt = datetime.fromisoformat(end_date).date()
    delta = (end_dt - TODAY).days
    if delta < 0:
        return "vencido"
    if delta <= 7:
        return "critico"
    if delta <= 15:
        return "atencion"
    return "ok"


def derive_days_remaining(end_date: str, status: str) -> int | None:
    if not end_date or "completado" in status.lower():
        return None
    end_dt = datetime.fromisoformat(end_date).date()
    return (end_dt - TODAY).days


def derive_level(edt: str) -> int:
    if not edt:
        return 0
    return edt.count(".") + 1


def derive_group(edt: str) -> str:
    if not edt:
        return "Sin grupo"
    return edt.split(".")[0]


def build_summary(record: dict[str, Any]) -> str:
    pieces = [
        f"EDT {record['edt']}" if record.get("edt") else "",
        to_text(record.get("responsable")),
        to_text(record.get("seguimiento_dgppcs")),
        to_text(record.get("estado")),
    ]
    return " | ".join(piece for piece in pieces if piece)


def refresh_hierarchy_fields(record: dict[str, Any]) -> None:
    edt = to_text(record.get("edt"))
    record["edt"] = edt
    record["nivel"] = derive_level(edt)
    record["grupo"] = derive_group(edt)
    record["resumen"] = build_summary(record)


def apply_canonical_hierarchy_corrections(records: list[dict[str, Any]]) -> None:
    for record in records:
        record_id = to_text(record.get("id"))
        actividad = to_text(record.get("actividad"))
        edt = to_text(record.get("edt"))
        if edt == "3" and re.search(r"otras condiciones solicitadas por el bm", actividad, re.IGNORECASE):
            record["actividad"] = CANONICAL_ROOT_3_TITLE
        if re.search(r"inicio de efectividad del pr[eé]stamo", actividad, re.IGNORECASE):
            record["actividad"] = CANONICAL_ROOT_4_TITLE
        if re.search(r"conformaci[oó]n equipo etgp.*dgppcs", actividad, re.IGNORECASE):
            record["actividad"] = CANONICAL_ETGP_TITLE
        if record_id in CANONICAL_EDT_BY_ID:
            record["edt"] = CANONICAL_EDT_BY_ID[record_id]
        if re.search(r"conformaci[oó]n equipo etgp.*dgppcs", actividad, re.IGNORECASE):
            record["edt"] = CANONICAL_EDT_BY_ID.get(record_id, record.get("edt"))
        if re.search(r"elaboraci[oó]n del sistema de seguimiento", actividad, re.IGNORECASE):
            record["edt"] = ""
        refresh_hierarchy_fields(record)


def build_record(raw: dict[str, Any], idx: int) -> dict[str, Any]:
    edt = to_text(raw.get("EDT"))
    actividad = to_text(raw.get("Actividad"))
    responsable = to_text(raw.get("Responsable"))
    inicio = to_date(raw.get("Inicio"))
    final = to_date(raw.get("Final"))
    estado = clean_status(raw.get("Estado"))
    seguimiento = to_text(raw.get("Responsable de Seguimiento DGPPCS"))
    contacto = to_text(raw.get("Contacto"))

    return {
        "id": f"case-{idx:03d}",
        "edt": edt,
        "actividad": actividad,
        "responsable": responsable,
        "inicio": inicio,
        "final": final,
        "duracion_cal": to_number(raw.get("Duración (Días Calendario)")),
        "dias_habiles": to_number(raw.get("Días Hábiless")),
        "dias_restantes": derive_days_remaining(final, estado),
        "estado": estado,
        "seguimiento_dgppcs": seguimiento,
        "contacto": contacto,
        "nivel": derive_level(edt),
        "grupo": derive_group(edt),
        "avance_estimado": derive_progress(estado),
        "alerta": derive_alert(inicio, final, estado),
        "resumen": " | ".join(
            [
                piece
                for piece in [
                    f"EDT {edt}" if edt else "",
                    responsable,
                    seguimiento,
                    estado,
                ]
                if piece
            ]
        ),
    }


def main() -> None:
    wb = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]

    headers = {
        col_idx: normalize_header(ws.cell(HEADER_ROW, col_idx).value, col_idx)
        for col_idx in SELECTED_COLUMNS
    }

    records: list[dict[str, Any]] = []
    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        row_values = {col_idx: ws.cell(row_idx, col_idx).value for col_idx in SELECTED_COLUMNS}
        if not any(value not in (None, "") for value in row_values.values()):
            continue

        row = {headers[col_idx]: row_values[col_idx] for col_idx in SELECTED_COLUMNS}
        actividad = to_text(row.get("Actividad"))
        edt = to_text(row.get("EDT"))
        if not actividad and not edt:
            continue
        records.append(build_record(row, len(records) + 1))

    apply_canonical_hierarchy_corrections(records)

    status_counts = Counter(record["estado"] for record in records)
    alert_counts = Counter(record["alerta"] for record in records)

    payload = {
        "meta": {
            "sourceFile": str(SOURCE_XLSX),
            "sheetName": SHEET_NAME,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "referenceDate": TODAY.isoformat(),
            "recordCount": len(records),
        },
        "summary": {
            "total": len(records),
            "completed": status_counts.get("Completado", 0),
            "inProgress": status_counts.get("En Proceso", 0),
            "pending": status_counts.get("Pendiente de Iniciar", 0),
            "withoutStatus": status_counts.get("Sin estado", 0),
            "critical": alert_counts.get("critico", 0),
            "overdue": alert_counts.get("vencido", 0),
            "attention": alert_counts.get("atencion", 0),
        },
        "records": records,
    }

    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_JSON.write_text(json_text, encoding="utf-8")
    OUTPUT_JS.write_text(f"window.PEC_TRACKING_DATA = {json_text};\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON} with {len(records)} records")
    print(f"Wrote {OUTPUT_JS}")


if __name__ == "__main__":
    main()
